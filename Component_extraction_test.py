import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

output_file = "BacDive_ChEBI_Gujer_Metabolite_Library_Option1_Workflow.xlsx"

# Sheet 1: existing curated metabolite library
metabolites = [
    {
        "IndividualName": "Acetate",
        "ChEBI_ID": "CHEBI:30089",
        "FunctionalGroup": "Electron donor / Carbon source",
        "ComponentClass": "SolubleComponent",
        "GujerNotation": "S_Ac",
        "EmpiricalFormula": "C2H3O2-",
        "Definition": "A soluble short-chain fatty acid used as a substrate in biological wastewater processes.",
        "Notes": "Relevant to acetogenesis, denitrification and sulfate reduction."
    },
    {
        "IndividualName": "Sulfate",
        "ChEBI_ID": "CHEBI:16189",
        "FunctionalGroup": "Electron acceptor / Sulfur species",
        "ComponentClass": "SolubleComponent",
        "GujerNotation": "S_SO4",
        "EmpiricalFormula": "SO4(2-)",
        "Definition": "A soluble sulfur oxyanion used as an electron acceptor by sulfate-reducing microorganisms.",
        "Notes": "Important in SANI and sulfate reduction models."
    },
    {
        "IndividualName": "Methane",
        "ChEBI_ID": "CHEBI:16183",
        "FunctionalGroup": "End product",
        "ComponentClass": "GasComponent",
        "GujerNotation": "G_CH4",
        "EmpiricalFormula": "CH4",
        "Definition": "A gaseous end product of methanogenic metabolism.",
        "Notes": "May also occur as dissolved methane depending on model structure."
    }
]

df_metabolites = pd.DataFrame(metabolites)

# Sheet 2: BacDive API fields to extract
bacdive_fields = pd.DataFrame([
    {
        "BacDiveSection": "metabolite_utilization",
        "ExpectedContent": "Metabolites used or consumed by microorganisms",
        "OntologyUse": "Electron donors, carbon sources, substrates"
    },
    {
        "BacDiveSection": "metabolite_production",
        "ExpectedContent": "Metabolites produced by microorganisms",
        "OntologyUse": "Products, intermediates, end products"
    },
    {
        "BacDiveSection": "chebi_ID",
        "ExpectedContent": "ChEBI identifier associated with BacDive metabolite",
        "OntologyUse": "External ontology reference"
    }
])

# Sheet 3: Option 1 workflow
workflow = pd.DataFrame([
    {"Step": 1, "Action": "Retrieve all BacDive strain IDs."},
    {"Step": 2, "Action": "Fetch each strain record using the BacDive API."},
    {"Step": 3, "Action": "Extract metabolite utilization and production fields."},
    {"Step": 4, "Action": "Extract metabolite names and ChEBI IDs where available."},
    {"Step": 5, "Action": "Deduplicate metabolites by name and ChEBI ID."},
    {"Step": 6, "Action": "Classify each metabolite as SolubleComponent, ParticulateComponent, or GasComponent."},
    {"Step": 7, "Action": "Add Gujer notation, empirical formula and definition."},
    {"Step": 8, "Action": "Export to Excel for ontology population."}
])

# Sheet 4: API extraction code
script_text = """
import requests
import pandas as pd

BASE_URL = "https://api.bacdive.dsmz.de"

def classify_component(name):
    gas_keywords = ["methane", "hydrogen gas", "oxygen", "carbon dioxide", "nitrogen gas"]
    particulate_keywords = ["cellulose", "starch", "elemental sulfur", "glycogen", "polyphosphate"]

    name_lower = name.lower()

    if any(k in name_lower for k in gas_keywords):
        return "GasComponent"
    elif any(k in name_lower for k in particulate_keywords):
        return "ParticulateComponent"
    else:
        return "SolubleComponent"

def suggest_gujer_notation(name, component_class):
    clean = name.replace(" ", "_").replace("-", "_")

    if component_class == "SolubleComponent":
        return f"S_{clean}"
    elif component_class == "ParticulateComponent":
        return f"X_{clean}"
    elif component_class == "GasComponent":
        return f"G_{clean}"
    else:
        return ""

# Pseudocode: replace with authenticated BacDive API calls
records = []

# Example structure after API extraction:
# for strain_id in strain_ids:
#     data = fetch_bacdive_record(strain_id)
#     for metabolite in extracted_metabolites:
#         records.append({
#             "IndividualName": metabolite["name"],
#             "ChEBI_ID": metabolite.get("chebi_ID", ""),
#             "FunctionalGroup": metabolite.get("source_field", ""),
#             "ComponentClass": classify_component(metabolite["name"]),
#             "GujerNotation": suggest_gujer_notation(
#                 metabolite["name"],
#                 classify_component(metabolite["name"])
#             ),
#             "EmpiricalFormula": "",
#             "Definition": "",
#             "BacDiveStrainID": strain_id
#         })

df = pd.DataFrame(records)
df = df.drop_duplicates(subset=["IndividualName", "ChEBI_ID"])
df.to_excel("Full_BacDive_ChEBI_Gujer_Metabolites.xlsx", index=False)
"""

# Write workbook
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_metabolites.to_excel(writer, sheet_name="Metabolite_Library", index=False)
    bacdive_fields.to_excel(writer, sheet_name="BacDive_API_Fields", index=False)
    workflow.to_excel(writer, sheet_name="Option1_Workflow", index=False)

    script_df = pd.DataFrame({"Python_Code": script_text.splitlines()})
    script_df.to_excel(writer, sheet_name="Python_Extraction_Script", index=False)

# Format workbook
from openpyxl import load_workbook

wb = load_workbook(output_file)

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 3, 60)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(output_file)

print(f"Workbook saved as: {output_file}")